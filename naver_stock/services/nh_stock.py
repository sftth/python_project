import requests
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from naver_stock.utils.file_utils import generate_dated_excel_filename, save_or_append_excel, save_excel
from naver_stock.utils.api_utils import append_data

# ✅ 조회할 종목 리스트
stock_list = [
["stock","HPSP","403870"],
["etf","KODEX 미디어&엔터테인먼트","266360"],
["etf","KODEX 코스닥150레버리지","233740"],
["stock","LG생활건강","051900"],
["stock","LG씨엔에스","064400"],
["stock","LG전자","066570"],
["stock","LS에코에너지","229640"],
["stock","Plus 한화그룹주","0000J0"],
["stock","POSCO홀딩스","005490"],
["stock","SAMG엔터","419530"],
["stock","SK하이닉스","000660"],
["stock","SOL 조선 TOP3플러스","466920"],
["etf","TIGER 200에너지화학레버리지","243890"],
["etf","TIGER 미국S&P500","360750"],
["etf","TIGER 미국나스닥100커버드콜(합성)","441680"],
["stock","대한항공","003490"],
["stock","두산에너빌리티","034020"],
["stock","디어유","376300"],
["stock","삼성전자","005930"],
["stock","삼성중공업","010140"],
["stock","세아제강","306200"],
["stock","솔트룩스","304100"],
["stock","쏠리드","050890"],
["stock","에코프로비엠","247540"],
["stock","오뚜기","007310"],
["stock","코오롱티슈진","950160"],
["stock","포스코인터내셔널","047050"],
["stock","포스코퓨처엠","003670"],
["stock","한국항공우주","047810"],
["stock","한샘","009240"],
["stock","한화에어로스페이스","012450"],
["stock","한화오션","042660"],
["stock","현대ADM","187660"],
["stock","현대로템","064350"],
["stock","현대지에프홀딩스","005440"],
["stock","현대차","005380"],
["stock","후성","093370"]
]

OUTPUT_DIR = "output/nh"

# ✅ API 기본 URL
base_url = "https://m.stock.naver.com/api/stock/{}/price?pageSize=1&page=1"



def nh_save_xlsx():
    # ✅ 데이터 저장할 리스트 생성
    all_stock_data = []

    # ✅ 각 종목별 데이터 요청
    headers = {"User-Agent": "Mozilla/5.0"}
    for category, name, stock_code in stock_list:
        url = base_url.format(stock_code)
        response = requests.get(url, headers=headers)

        if response.status_code == 200:
            data = response.json()
            all_stock_data += append_data(data, firm="NH", category=category, name=name, stock_code=stock_code)
        else:
            print(f"⛔ {stock_code} 데이터 가져오기 실패 (HTTP {response.status_code})")

    # ✅ 데이터프레임 변환
    df = pd.DataFrame(all_stock_data)

    # ✅ 엑셀 파일로 저장
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    excel_filename = generate_dated_excel_filename(prefix="nh", output_dir=OUTPUT_DIR)
    print(f"⛔Excel file name: {excel_filename} ")
#    excel_filename = "nh_stocks_data_v1.1.xlsx"
#    df.to_excel(excel_filename, index=False, sheet_name="Stock Prices")
#    save_or_append_excel(df, excel_filename, sheet_name="Stock Prices")
    save_excel(df, excel_filename, sheet_name="Stock Prices")
    # ✅ 엑셀 파일 불러와 서식 적용
    wb = load_workbook(excel_filename)
    ws = wb["Stock Prices"]

    # ✅ 스타일 설정
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")  # 제목 폰트
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # 제목 배경색
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center", vertical="center")

    # ✅ 숫자 스타일 지정
    currency_style = NamedStyle(name="currency_style", number_format=r'_(₩* #,##0_);_(₩* (#,##0);_(₩* "-"_);_(@_)')
    percent_style = NamedStyle(name="percent_style", number_format="0.00%")
    date_style = NamedStyle(name="date_style", number_format="yyyy.m.d")

    # ✅ 스타일을 워크북에 추가 (기존 스타일 중복 방지)
    if "currency_style" not in wb.named_styles:
        wb.add_named_style(currency_style)
    if "percent_style" not in wb.named_styles:
        wb.add_named_style(percent_style)
    if "date_style" not in wb.named_styles:
        wb.add_named_style(date_style)

    # ✅ 헤더 스타일 적용
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # ✅ 데이터 셀 스타일 적용
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[5].style == "Normal":  # 날짜 셀
            row[5].style = date_style
        if row[6].style == "Normal":  # 종가
            row[6].style = currency_style
        if row[7].style == "Normal":  # 전일 대비
            row[7].style = currency_style
        if row[8].style == "Normal":  # 등락률
            row[8].style = percent_style
        if row[9].style == "Normal":  # 시가
            row[9].style = currency_style
        if row[10].style == "Normal":  # 고가
            row[10].style = currency_style
        if row[11].style == "Normal":  # 저가
            row[11].style = currency_style

    # ✅ Saving excel file
    wb.save(excel_filename)
    print(f"✅ Saving excel file is completed successfully: {excel_filename}")
