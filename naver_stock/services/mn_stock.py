import requests
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from naver_stock.utils.file_utils import generate_dated_excel_filename, save_or_append_excel, save_excel
from naver_stock.utils.api_utils import append_data

# ✅ 조회할 종목 리스트
stock_list = [
["stock","애플","AAPL.O"],
["stock","아마존","AMZN.O"],
["stock","뱅크오브아메리카","BAC"],
["stock","블랙록","BLK"],
["stock","블랙스톤","BX"],
["stock","코인베이스 글로벌","COIN.O"],
["stock","세일즈포스닷컴","CRM"],
["etf","SPDR 다우존스 ETF","DIA"],
["stock","알파벳C주","GOOG.O"],
["stock","골드만 삭스 그룹","GS"],
["stock","아이온큐","IONQ.K"],
["etf","아이셰어즈 코어 S&P500 ETF","IVV"],
["etf","아이셰어즈 글로벌 금융 ETF","IXG"],
["etf","JP모간 에쿼티 프리미엄 인컴 ETF","JEPI.K"],
["stock","제이피모간 체이스","JPM"],
["stock","코카콜라","KO"],
["stock","메타 플랫폼스","META.O"],
["stock","마이크로소프트","MSFT.O"],
["stock","넷플릭스","NFLX.O"],
["stock","엔비디아","NVDA.O"],
["stock","리얼티 인컴","O"],
["stock","유니버설 디스플레이","OLED.O"],
["etf","아이셰어즈 우선주 ETF","PFF.O"],
["stock","팔란티어 테크놀로지스 A주","PLTR.O"],
["stock","펠로톤 인터렉티브","PTON.O"],
["etf","인베스코 QQQ ETF","QQQ.O"],
["etf","인베스코 나스닥100 ETF","QQQM.O"],
["etf","글로벌엑스 나스닥 100 커버드콜 ETF","QYLD.O"],
["etf","슈왑 미국 배당주 ETF","SCHD.K"],
["stock","슈퍼 마이크로 컴퓨터","SMCI.O"],
["stock","소파이 테크놀로지스","SOFI.O"],
["etf","인베스코 S&P500 고배당 저변동성 ETF","SPHD.K"],
["etf","인베스코 S&P500 저변동성 ETF","SPLV.K"],
["etf","SPDR S&P500 ETF","SPY"],
["etf","프로셰어즈 나스닥100 3배 레버리지 ETF","TQQQ.O"],
["stock","테슬라","TSLA.O"],
["stock","유나이티드 에어라인 홀딩스","UAL.O"],
["stock","업스타트 홀딩스","UPST.O"],
["etf","뱅가드 S&P500 ETF","VOO"],
["etf","SPDR 미국 항공우주&방위산업 ETF","XAR"],
["stock","블록","XYZ"]
]

OUTPUT_DIR = "output/mn"

# ✅ API 기본 URL
base_url = "https://api.stock.naver.com/{}/{}/price?page=1&pageSize=1"

def mn_save_xlsx():
    # ✅ 데이터 저장할 리스트 생성
    all_stock_data = []
    # ✅ 각 종목별 데이터 요청
    headers = {"User-Agent": "Mozilla/5.0"}
    for category, name, stock_code in stock_list:
        url = base_url.format(category, stock_code)
        response = requests.get(url, headers=headers)

        if response.status_code == 200:
            data = response.json()
            all_stock_data += append_data(data, firm="MN", category=category, name=name, stock_code=stock_code)

        else:
            print(f"⛔ {category} {stock_code} 데이터 가져오기 실패 (HTTP {response.status_code})")

    # ✅ 데이터프레임 변환
    df = pd.DataFrame(all_stock_data)

    # ✅ 엑셀 파일로 저장
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    excel_filename = generate_dated_excel_filename(prefix="mn", output_dir=OUTPUT_DIR)
#    excel_filename = "mn_stocks_data_v1.1.xlsx"
#    df.to_excel(excel_filename, index=False, sheet_name="Stock Prices")

    #save_or_append_excel(df, excel_filename, sheet_name="Stock Prices")
    save_excel(df, excel_filename, sheet_name="Stock Prices")

    # ✅ 엑셀 파일 불러와 서식 적용
    wb = load_workbook(excel_filename)
    ws = wb["Stock Prices"]

    # ✅ 스타일 설정
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")  # 제목 폰트
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # 제목 배경색
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center", vertical="center")

    # ✅ 숫자 스타일 지정
    currency_style = NamedStyle(name="currency_style", number_format=r'_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)')
    percent_style = NamedStyle(name="percent_style", number_format="0.00%")
    date_style = NamedStyle(name="date_style", number_format="yyyy.m.d")

    # ✅ 스타일을 워크북에 추가 (기존 스타일 중복 방지)
    if "currency_style" not in wb.named_styles:
        wb.add_named_style(currency_style)
    if "percent_style" not in wb.named_styles:
        wb.add_named_style(percent_style)
    if "date_style" not in wb.named_styles:
        wb.add_named_style(date_style)

    # ✅ 헤더 스타일 적용
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # ✅ 데이터 셀 스타일 적용
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[5].style == "Normal":  # 날짜 셀
            row[5].style = date_style
        if row[6].style == "Normal":  # 종가
            row[6].style = currency_style
        if row[7].style == "Normal":  # 전일 대비
            row[7].style = currency_style
        if row[8].style == "Normal":  # 등락률
            row[8].style = percent_style
        if row[9].style == "Normal":  # 시가
            row[9].style = currency_style
        if row[10].style == "Normal":  # 고가
            row[10].style = currency_style
        if row[11].style == "Normal":  # 저가
            row[11].style = currency_style

    # ✅ 엑셀 파일 저장
    wb.save(excel_filename)
    print(f"✅ 모든 종목 데이터가 엑셀에 저장 완료 (서식 적용): {excel_filename}")
