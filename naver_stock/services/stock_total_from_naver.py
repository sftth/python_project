from http.client import responses
from unicodedata import category
import platform
from datetime import datetime
import requests
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from naver_stock.utils.file_utils import generate_dated_excel_filename, save_excel
from naver_stock.utils.api_utils import append_data

# ✅ 조회할 종목 리스트

OUTPUT_DIR = "output/nh"

# API 기본 URL
realtime_url = "https://polling.finance.naver.com/api/realtime/{}/stock/{}" #stockName, symbolCode, closePrice, hightPrice, lowPrice, openPrice
domestic_url = "https://m.stock.naver.com/api/stock/{}/integration"

worldstock_basic_url = "https://api.stock.naver.com/stock/{}/basic"
worldstock_url = "https://api.stock.naver.com/stock/{}/integration"

# 데이터 저장할 리스트 생성
all_stock_data = []

# 코드 매핑: (엑셀열이름, 소수점변환여부)
target_codes = {
    "per": ("PER", True),
    "eps": ("EPS", True),
    "pbr": ("PBR", True),
    "bps": ("BPS", True),
    "dividendYieldRatio": ("배당수익률", False),
    "dividend": ("주당배당금", True),
    "cnsPer": ("추정PER", True),
    "cnsEps": ("추정EPS", True)
}

# 함수 선언
def clean_and_convert(value: str, to_percent: bool) -> float:
    # "6.04배" → 6.04, "3.48%" → 3.48, "750원" → 750
    # print("value:" + value)
    try:
        if to_percent:
            num = value.replace(",", "").replace("배", "").replace("원", "")
            number = float(num)
        else:
            number = value

        # print(f"원본 값: {value} -> 변환 후: {number}")  
        return number

    except Exception as e:
        # print(f"값 변환 오류(value: {value}) -> {e}")
        return None

def get_stock_info(stock_list: list, prefix="nh", output_dir="/output/"):
    # 데이터 저장용 리스트 생성
    all_stock_data = []
    # 각 종목별 데이터 요청
    headers = {"User-Agent": "Mozilla/5.0"}

    for market, category, name, stock_code in stock_list:
        stock_info = []
        # RealTime Info URL
        url = realtime_url.format(market, stock_code)

        response = requests.get(url, headers=headers)

        if response.status_code == 200:
            data = response.json()
            stock_info_list = get_realtime_data(data['datas'], firm=prefix.upper(), category=category, name=name, stock_code=stock_code)

        all_stock_data += additional_infos(stock_info_list, stock_code, market=market)

    save_info_to_excel(all_stock_data, prefix=prefix, output_dir=output_dir)

def get_realtime_data(data, firm, category, name, stock_code):
    # 운영체제에 따라 날짜 포맷 선택
    if platform.system() == "Windows":
        date_format = "%Y.%#m.%#d"
    else:  # macOS, Linux
        date_format = "%Y.%-m.%-d"

    realtime_data = []
    if isinstance(data, list):
        for item in data:
            realtime_data.append({
                "NO": "",
                "Firm": firm,
                "Type": category,
                "Name": name,
                "Code": stock_code,
                "Date": datetime.strptime(item["localTradedAt"][:10], "%Y-%m-%d"),
                "Price": float(item["closePrice"].replace(",", "")),  # 쉼표 제거 후 정수 변환
                "Change": float(item["compareToPreviousClosePrice"].replace(",", "")),  # ✅ 쉼표 제거 후 int 변환
                "Rate (%)": float(item["fluctuationsRatio"]) * 0.01,  # 등락률은 소수점 필요
                "Open": float(item["openPrice"].replace(",", "")),
                "High": float(item["highPrice"].replace(",", "")),
                "Low": float(item["lowPrice"].replace(",", "")),
                "주식수":"",
                "매수단가":"",
                "매수금액":"",
                "평가금액":"",
                "이익":"",
                "이익률":""
            })
    return realtime_data


def additional_infos(stock_info_list,stock_code, market="domestic"):
    # 각 종목별 데이터 요청
    headers = {"User-Agent": "Mozilla/5.0"}
    try:
        # Assign url for domestic or worldstock
        if market == 'domestic':
            url = domestic_url.format(stock_code)

            response = requests.get(url, headers=headers)

            if response.status_code == 200:
                data = response.json()

                for item in data.get("totalInfos", []):
                    if item["code"] in target_codes:
                        label, need_float = target_codes[item["code"]]
                        value = clean_and_convert(item["value"], need_float)

                        # 모든 딕셔너리에 추가
                        for stock_info in stock_info_list:
                            stock_info[label] = value
                try:
                    if data['consensusInfo'] is not None:
                        price_target_str = data["consensusInfo"]["priceTargetMean"]
                        # print(f"목표주: {price_target_str}")
                        price_target = float(price_target_str.replace(",", ""))
                        for stock_info in stock_info_list:
                            stock_info["목표주가"] = price_target
                except Exception as e:
                    print(f"값 변환 오류(priceTargetMean: {price_target_str}) -> {e}")
        else:
            url_basic = worldstock_basic_url.format(stock_code)
            url_additional = worldstock_url.format(stock_code)

            response_basic = requests.get(url_basic, headers=headers)
            response_additional = requests.get(url_additional, headers=headers)

            if response_basic.status_code == 200:
                data = response_basic.json()

                for item in data.get("stockItemTotalInfos", []):
                    if item["code"] in target_codes:
                        label, need_float = target_codes[item["code"]]
                        value = clean_and_convert(item["value"], need_float)

                        # 모든 딕셔너리에 추가
                        for stock_info in stock_info_list:
                            stock_info[label] = value

            if response_additional.status_code == 200:
                data = response_additional.json()
                if data['consensusInfo'] is not None:
                    price_target_str = data["consensusInfo"]["priceTargetMean"]
                    # print(f"목표주: {price_target_str}")
                    price_target = float(price_target_str.replace(",", ""))
                    for stock_info in stock_info_list:
                        stock_info["목표주가"] = price_target
    except Exception as e:
        print(f"값 변환 오류(priceTargetMean: {price_target_str}) -> {e}")

    return stock_info_list

def save_info_to_excel(all_stock_data, prefix="nh", output_dir="/output"):
    # 데이터프레임 변환
    df = pd.DataFrame(all_stock_data)

    if df.empty:
        print("데이터가 없습니다. API 응답을 확인하세요.")
    #else:
    #    print(df.head()) # 데이터 확인

    # 엑셀 파일로 저장
    #os.makedirs(OUTPUT_DIR, exist_ok=True)

    excel_filename = generate_dated_excel_filename(prefix=prefix, output_dir=output_dir)
    save_excel(df, excel_filename, sheet_name="Stock TotalInfos")
    #excel_filename = "nh_total_stocks_data_v1.0.xlsx"
    #df.to_excel(excel_filename, index=False, sheet_name="Stock TotalInfos")

    # ✅ 엑셀 파일 불러와 서식 적용
    wb = load_workbook(excel_filename)
    ws = wb["Stock TotalInfos"]

    # ✅ 스타일 설정
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")  # 제목 폰트
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # 제목 배경색
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # ✅ 숫자 스타일 지정
    currency_style = NamedStyle(name="currency_style", number_format=r'_(₩* #,##0_);_(₩* (#,##0);_(₩* "-"_);_(@_)')
    number_style = NamedStyle(name="number_style",  number_format="#,##0.00")
    percent_style = NamedStyle(name="percent_style", number_format="0.00%")
    date_style = NamedStyle(name="date_style", number_format="yyyy.mm.dd")

    # ✅ 스타일을 워크북에 추가 (기존 스타일 중복 방지)
    if "currency_style" not in wb.named_styles:
        wb.add_named_style(currency_style)
    if "percent_style" not in wb.named_styles:
        wb.add_named_style(percent_style)
    if "date_style" not in wb.named_styles:
        wb.add_named_style(date_style)

    # ✅ 헤더 스타일 적용
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # ✅ 데이터 셀 스타일 적용
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        #row[5].style = currency_style  # PBR float 형식 적용
        #row[6].style = currency_style  # PBR float 형식 적용
        row[0]
        row[1]
        row[2]
        row[3].number_format = '#,##0'
        row[4]
        row[5].style = date_style
        row[6].style = number_style
        row[7].style = number_style
        row[8].style = percent_style
        row[9].style = number_style
        row[10].style = number_style
        row[11].style = number_style
        row[12].style = number_style
        row[13].style = number_style
        row[14].style = number_style
        row[15].style = number_style
        row[16].style = percent_style
        row[17].style = number_style
        row[18].style = number_style
        row[19].style = number_style
        row[20].style = number_style
        row[21].style = number_style
        row[22].style = number_style
        row[23].style = number_style
        row[24].style = number_style
    # ✅ 엑셀 파일 저장
    wb.save(excel_filename)
    print(f"✅ 모든 종목 데이터가 엑셀에 저장 완료 (서식 적용): {excel_filename}")
