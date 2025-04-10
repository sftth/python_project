import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from datetime import datetime

# ✅ 조회할 종목 리스트
stock_list = [
["대한항공","003490"],
["두산에너빌리티","034020"],
["디어유","376300"],
["삼성전자","005930"],
["삼성중공업","010140"],
["쏠리드","050890"],
["에코프로비엠","247540"],
["코오롱티슈진","950160"],
["포스코인터내셔널","047050"],
["포스코퓨처엠","003670"],
["한국항공우주","047810"],
["한화솔루션","009835"],
["한화에어로스페이스","012450"],
["한화오션","042660"],
["현대로템","064350"],
["현대차","005380"],
["현대ADM","187660"],
["후성","093370"],
["HPSP","403870"],
["KODEX 미디어&엔터테인먼트","266360"],
["KODEX 코스닥150레버리지","233740"],
["LG생활건강","051900"],
["LG씨엔에스","064400"],
["LG전자","066570"],
["LS에코에너지","229640"],
["Plus 한화그룹주","0000J0"],
["POSCO홀딩스","005490"],
["SAMG엔터","419530"],
["SK하이닉스","000660"],
["SOL 조선 TOP3플러스","466920"],
["TIGER 200에너지화학레버리지","243890"],
["TIGER 미국나스닥100커버드콜(합성)","441680"],
["TIGER 미국S&P500","360750"]
]

# API 기본 URL
base_url = "https://m.stock.naver.com/api/stock/{}/integration"

# 데이터 저장할 리스트 생성
all_stock_data = []

# 코드 매핑: (엑셀열이름, 소수점변환여부)
target_codes = {
    "per": ("PER", True), 
    "pbr": ("PBR", True),
    "dividendYieldRatio": ("배당수익률", False),
    "dividend": ("주당배당금", True)
}

# 함수 선언
def clean_and_convert(value: str, to_percent: bool) -> float:
    # "6.04배" → 6.04, "3.48%" → 3.48, "750원" → 750
    print("value:" + value)
    try:
        if to_percent:
            num = value.replace(",", "").replace("배", "").replace("원", "")
            number = float(num)
        else:
            number = value

        # print(f"원본 값: {value} -> 변환 후: {number}")  
        return number

    except Exception as e:
        print(f"값 변환 오류(value: {value}) -> {e}")
        return None


# ✅ 각 종목별 데이터 요청
headers = {"User-Agent": "Mozilla/5.0"}

for name, stock_code in stock_list:
    url = base_url.format(stock_code)

    response = requests.get(url, headers=headers)

    if response.status_code == 200:
        data = response.json()

        stock_info = {
            "종목명": name,
            "종목코드": stock_code
        }    

        for item in data.get("totalInfos", []):
            if item["code"] in target_codes:
                label, need_float = target_codes[item["code"]]
                value = clean_and_convert(item["value"], need_float)

                stock_info[label] = value

        try:
            price_target_str = data["consensusInfo"]["priceTargetMean"]
            print(f"목표주: {price_target_str}")
            stock_info["목표주가"] = float(price_target_str.replace(",", ""))
        except Exception as e:
            print(f"값 변환 오류(priceTargetMean: {price_target_str}) -> {e}")

        all_stock_data.append(stock_info)
    else:
        print(f"⛔ {stock_code} 데이터 가져오기 실패 (HTTP {response.status_code})")

# ✅ 데이터프레임 변환
df = pd.DataFrame(all_stock_data)

if df.empty:
    print("데이터가 없습니다. API 응답을 확인하세요.")
#else:
#    print(df.head()) # 데이터 확인 

# ✅ 엑셀 파일로 저장
excel_filename = "nh_total_stocks_data_v1.0.xlsx"
df.to_excel(excel_filename, index=False, sheet_name="Stock TotalInfos")

# ✅ 엑셀 파일 불러와 서식 적용
wb = load_workbook(excel_filename)
ws = wb["Stock TotalInfos"]

# ✅ 스타일 설정
header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")  # 제목 폰트
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # 제목 배경색
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
center_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")

# ✅ 숫자 스타일 지정
currency_style = NamedStyle(name="currency_style", number_format=r'_(₩* #,##0_);_(₩* (#,##0);_(₩* "-"_);_(@_)')
percent_style = NamedStyle(name="percent_style", number_format="0.00%")
date_style = NamedStyle(name="date_style", number_format="yyyy.mm.dd")

# ✅ 스타일을 워크북에 추가 (기존 스타일 중복 방지)
if "currency_style" not in wb.named_styles:
    wb.add_named_style(currency_style)
if "percent_style" not in wb.named_styles:
    wb.add_named_style(percent_style)
if "date_style" not in wb.named_styles:
    wb.add_named_style(date_style)

# ✅ 헤더 스타일 적용
for cell in ws[1]:  
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# ✅ 데이터 셀 스타일 적용
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    row[4].alignment = left_align
    row[5].style = currency_style  # PBR float 형식 적용
    row[6].style = currency_style  # PBR float 형식 적용

# ✅ 엑셀 파일 저장
wb.save(excel_filename)
print(f"✅ 모든 종목 데이터가 엑셀에 저장 완료 (서식 적용): {excel_filename}")
