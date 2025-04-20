import requests
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from naver_stock.utils.file_utils import generate_dated_excel_filename, save_or_append_excel, save_excel
from naver_stock.utils.api_utils import append_data

# ✅ 조회할 종목 리스트
stock_list = [
["stock","그래디언트","035080"],
["stock","대원미디어","048910"],
["stock","에스디바이오센서","137310"],
["stock","웅진씽크빅","095720"],
["stock","카카오","035720"],
["stock","카카오게임즈","293490"],
["stock","카카오뱅크","323410"],
["stock","펄어비스","263750"],
["stock","LG에너지솔루션","373220"],
["stock","LG전자","066570"],
["stock","NAVER","035420"],
["stock","SK리츠","395400"],
["stock","TIGER 차이나전기차SOLACTIVE","371460"],
["stock","TIGER 차이나항셍테크","371160"],
["stock","TIGER배당커버드콜액티브","472150"],
["stock","TIGER은행고배당플러스 TOP10","466940"]
]

stock_list_us = [
["etf","YieldMax COIN Option Income Strategy ETF","CONY.K"],
["etf","Volatility shares bitcoin strategy 2x ETF","BITX.K"]
]

OUTPUT_DIR = "output"

# ✅ API 기본 URL
base_url = "https://m.stock.naver.com/api/{}/{}/price?pageSize=1&page=1"
base_url_us = "https://api.stock.naver.com/{}/{}/price?page=1&pageSize=1"

# ✅ 각 종목별 데이터 요청
headers = {"User-Agent": "Mozilla/5.0"}


def fetch_stock_data(stock_list, base_url):
    stock_data = []
    for category, name, stock_code in stock_list:
        url = base_url.format(category, stock_code)

        try:
            response = requests.get(url, headers=headers)
            if response.status_code == 200:
                data = response.json()
                stock_data += append_data(data, firm="HT", category=category, name=name, stock_code=stock_code)
            else:
                print(f"⛔ {stock_code} 데이터 가져오기 실패 (HTTP {response.status_code})")
        except requests.exceptions.RequestException as e:
            print(f"⛔ {stock_code} 데이터 가져오기 실패: {e}")
    return stock_data

def ht_save_xlsx():
    # KR & US stock data
    all_stock_data = fetch_stock_data(stock_list, base_url) + fetch_stock_data(stock_list_us, base_url_us)


    # ✅ 데이터프레임 변환
    df = pd.DataFrame(all_stock_data)

    if df.empty:
        print("⚠️ 데이터가 없습니다. API 응답을 확인하세요.")
#    else:
#        print(df.head())  # ✅ 데이터가 들어있는지 확인

    # ✅ 엑셀 파일로 저장
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    excel_filename = generate_dated_excel_filename(prefix="ht", output_dir=OUTPUT_DIR)
    print(f"⛔Excel file name: {excel_filename} ")

    #save_or_append_excel(df, excel_filename, sheet_name="Stock Prices")
    save_excel(df, excel_filename, sheet_name="Stock Prices")

    # ✅ 엑셀 파일 불러와 서식 적용
    wb = load_workbook(excel_filename)
    ws = wb["Stock Prices"]

    # ✅ 스타일 설정
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")  # 제목 폰트
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # 제목 배경색
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center", vertical="center")

    # ✅ 숫자 스타일 지정
    currency_style = NamedStyle(name="currency_style", number_format="#,##0")
    percent_style = NamedStyle(name="percent_style", number_format="0.00%")
    date_style = NamedStyle(name="date_style", number_format="yyyy.m.d")

    # ✅ 스타일을 워크북에 추가 (기존 스타일 중복 방지)
    if "currency_style" not in wb.named_styles:
        wb.add_named_style(currency_style)
    if "percent_style" not in wb.named_styles:
        wb.add_named_style(percent_style)
    if "date_style" not in wb.named_styles:
        wb.add_named_style(date_style)

    # ✅ 헤더 스타일 적용
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # ✅ 데이터 셀 스타일 적용
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[5].style == "Normal":  # 날짜 셀
            row[5].style = date_style
        if row[6].style == "Normal":  # 종가
            row[6].style = currency_style
        if row[7].style == "Normal":  # 전일 대비
            row[7].style = currency_style
        if row[8].style == "Normal":  # 등락률
            row[8].style = percent_style
        if row[9].style == "Normal":  # 시가
            row[9].style = currency_style
        if row[10].style == "Normal":  # 고가
            row[10].style = currency_style
        if row[11].style == "Normal":  # 저가
            row[11].style = currency_style

    # ✅ 엑셀 파일 저장
    wb.save(excel_filename)
    wb.close() # 파일 닫기
    print(f"✅ 모든 종목 데이터가 엑셀에 저장 완료 (서식 적용): {excel_filename}")
