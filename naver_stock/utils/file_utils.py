import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from openpyxl import Workbook

def generate_dated_excel_filename(prefix="stock", output_dir="output"):
    """
    날짜 기반 파일명을 사용해 엑셀 파일을 저장하는 공통 함수.
    예: ht-2025.04.10.xlsx, nh-2025.04.10.xlsx

    :param data: dict 형태로 엑셀에 넣을 데이터 (key: 셀, value: 값)
    :param prefix: 파일명 접두어 ('ht', 'mn', 'nh' 등)
    :param output_dir: 저장 디렉토리 경로
    :return: 저장된 파일 경로 (str)
    """
    os.makedirs(output_dir, exist_ok=True)

    today_str = datetime.today().strftime("%Y.%m")
    filename = f"{prefix}-{today_str}.xlsx"
    return os.path.join(output_dir, filename)

def save_or_append_excel(df, output_path, sheet_name="Stock Prices"):
    #If there is a no directory, create it
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    if os.path.exists(output_path):
        try:
            # Load the excel file
            book = load_workbook(output_path)
            print(f"✅ Workbook loaded: {output_path}")
            print(f"📄 Sheet names: {book.sheetnames}")

            # Check sheet existence and row
            if sheet_name  in book.sheetnames:
                startrow = book[sheet_name].max_row
            else:
                startrow = 0

            print(f"✅ startrow: {startrow}")

            # 최신 pandas 방식으로 writer 열기
            with pd.ExcelWriter(
                output_path,
                engine="openpyxl",
                mode="a",
                if_sheet_exists="overlay"
            ) as writer:
                df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                    header=(startrow == 0),  # 첫 번째 행이면 헤더 포함
                    startrow=startrow
                )

            print("📌 데이터가 기존 시트에 추가되었습니다.")

        except Exception as e:
            print(f"❌ Error while loading or writing to Excel: {e}")
    else:
        # Create excel if there is not
        df.to_excel(output_path, index=False, sheet_name=sheet_name)
        print(f"📁 Created new Excel file: {output_path}")
